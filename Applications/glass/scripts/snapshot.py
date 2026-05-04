"""Glass build-state snapshot — conclusional artifact.

Run:  uv run --with openpyxl python scripts/snapshot.py
Out:  snapshots/glass-state-YYYY-MM-DD.json
      snapshots/glass-state-YYYY-MM-DD.xlsx
"""

from __future__ import annotations

import json
import os
import re
import sys
from datetime import date, timezone, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
BRIDGE = ROOT / "bridge"
SNAPSHOTS = ROOT / "snapshots"

# ── Manifest: every planned + emergent file, with metadata ────────────────────

MANIFEST: list[dict] = [
    # bridge
    {"path": "bridge/schema.ts", "category": "bridge", "designed": True},
    # main
    {"path": "src/main/index.ts", "category": "main", "designed": True},
    {"path": "src/main/bridge-watcher.ts", "category": "main", "designed": True},
    # preload
    {"path": "src/preload/index.ts", "category": "preload", "designed": True},
    # renderer entry
    {"path": "src/renderer/index.html", "category": "renderer", "designed": True},
    {"path": "src/renderer/index.ts", "category": "renderer", "designed": True},
    # field
    {"path": "src/renderer/field/Field.ts", "category": "field", "designed": True},
    {"path": "src/renderer/field/Presence.ts", "category": "field", "designed": True},
    {"path": "src/renderer/field/ThresholdLine.ts", "category": "field", "designed": True},
    {"path": "src/renderer/field/DiskEngine.ts", "category": "field", "designed": False},
    {"path": "src/renderer/field/OvalStadium.ts", "category": "field", "designed": False},
    {"path": "src/renderer/field/ModulationEngine.ts", "category": "field", "designed": False},
    # blocks
    {"path": "src/renderer/blocks/BlockManager.ts", "category": "blocks", "designed": True},
    {"path": "src/renderer/blocks/CodeBlock.ts", "category": "blocks", "designed": True},
    # conversation
    {"path": "src/renderer/conversation/ConversationLayer.ts", "category": "conversation", "designed": True},
    # state
    {"path": "src/renderer/state/FieldState.ts", "category": "state", "designed": True},
    {"path": "src/renderer/state/SessionState.ts", "category": "state", "designed": True},
    # assets
    {"path": "assets/spaceman.svg", "category": "assets", "designed": True},
    # config
    {"path": "electron.vite.config.ts", "category": "config", "designed": True},
    # emergent / unplanned
    {"path": "src/renderer/audio/AudioEngine.ts", "category": "audio", "designed": False},
]

EXPORT_RE = re.compile(
    r"export\s+(?:default\s+)?(?:class|interface|type|const|function|enum)\s+(\w+)"
)


def scan_file(abs_path: Path) -> dict:
    try:
        text = abs_path.read_text(encoding="utf-8")
    except Exception:
        return {"lines": 0, "exports": [], "mtime": None}
    lines = text.count("\n") + (1 if text and not text.endswith("\n") else 0)
    exports = EXPORT_RE.findall(text)
    mtime = datetime.fromtimestamp(abs_path.stat().st_mtime, tz=timezone.utc).isoformat()
    return {"lines": lines, "exports": exports, "mtime": mtime}


def build_inventory() -> list[dict]:
    rows = []
    for entry in MANIFEST:
        rel = entry["path"]
        abs_path = ROOT / rel
        exists = abs_path.exists()
        info = scan_file(abs_path) if exists else {"lines": 0, "exports": [], "mtime": None}

        status = "done" if exists and info["lines"] > 0 else "not_started"

        rows.append({
            "component": Path(rel).stem,
            "category": entry["category"],
            "path": rel,
            "lines": info["lines"],
            "status": status,
            "designed": entry["designed"],
            "exports": info["exports"],
            "mtime": info["mtime"],
        })
    return rows


def read_modulation_data() -> list[dict]:
    """Parse ENVELOPES and RECIPE from ModulationEngine.ts for the ceremony sheet."""
    mod_path = ROOT / "src/renderer/field/ModulationEngine.ts"
    if not mod_path.exists():
        return []

    text = mod_path.read_text(encoding="utf-8")

    env_re = re.compile(
        r'(\w+):\s*\{\s*sustain:\s*([\d.]+),\s*lfoRate:\s*([\d.]+),\s*lfoDepth:\s*([\d.]+)\s*\}'
    )
    envs = {m.group(1): {"sustain": float(m.group(2)), "lfoRate": float(m.group(3)), "lfoDepth": float(m.group(4))}
            for m in env_re.finditer(text)}

    base_disk_re = re.compile(r'disk:\s*\{\s*scale:\s*([\d.]+)')
    recipe_disk_re = re.compile(r'RECIPE[\s\S]*?disk:\s*\{\s*scale:\s*([\d.]+)')
    base_oval_re = re.compile(r'BASE[\s\S]*?oval:\s*\{[^}]*opacity:\s*([\d.]+)')
    recipe_oval_re = re.compile(r'RECIPE[\s\S]*?oval:\s*\{[^}]*opacity:\s*([\d.]+)')
    base_voice_re = re.compile(r'BASE[\s\S]*?voice:\s*\{[^}]*alpha:\s*([\d.]+)')
    recipe_voice_re = re.compile(r'RECIPE[\s\S]*?voice:\s*\{[^}]*alpha:\s*([\d.]+)')

    base_disk = float(m.group(1)) if (m := base_disk_re.search(text)) else 0
    max_disk = float(m.group(1)) if (m := recipe_disk_re.search(text)) else 0
    base_oval = float(m.group(1)) if (m := base_oval_re.search(text)) else 0
    max_oval = float(m.group(1)) if (m := recipe_oval_re.search(text)) else 0
    base_voice = float(m.group(1)) if (m := base_voice_re.search(text)) else 0
    max_voice = float(m.group(1)) if (m := recipe_voice_re.search(text)) else 0

    descriptions = {
        "ground": "Dim field, small disk, quiet ambient motion",
        "evaluating": "Cool-toned disk at ~50%, evaluation pulse ring",
        "floor_rising": "Disk scaling toward full, warm arc sweep on rim",
        "voices_appearing": "Three holographic figures fade in at field positions",
        "voice_1_active": "Voice I (amber/left) speaking, pulsing dot above helmet",
        "voice_2_active": "Voice II (silver/center) speaking",
        "voice_3_active": "Voice III (gold/right) speaking",
        "elevated": "Full brightness, all voices present, disk at max scale",
        "returning": "Field dimming, disk contracting, voices fading",
        "denied": "Brief flash, rapid LFO flicker, disk shrinks to minimum",
    }

    rows = []
    for state_name in [
        "ground", "evaluating", "floor_rising", "voices_appearing",
        "voice_1_active", "voice_2_active", "voice_3_active",
        "elevated", "returning", "denied",
    ]:
        env = envs.get(state_name, {"sustain": 0, "lfoRate": 0, "lfoDepth": 0})
        s = env["sustain"]
        rows.append({
            "state": state_name,
            "sustain": s,
            "lfo_rate": env["lfoRate"],
            "lfo_depth": env["lfoDepth"],
            "disk_scale": f"{base_disk + s * max_disk:.2f}",
            "oval_opacity": f"{base_oval + s * max_oval:.2f}",
            "voice_alpha": f"{base_voice + s * max_voice:.2f}",
            "description": descriptions.get(state_name, ""),
        })
    return rows


DEV_PHASES = [
    ("0", "Cleanup", "snapshot.py + xlsx", "-", "1h", "manual run", "pending"),
    ("0", "Cleanup", "npm run typecheck", "-", "15m", "CI gate", "pending"),
    ("1", "Morning", "Vitest setup", "-", "1h", "vitest.config.ts + first test", "not_started"),
    ("1", "Morning", "SessionState.ts", "FieldState.ts", "2h", "test: persist/restore round-trip", "not_started"),
    ("1", "Morning", "ThresholdLine.ts", "ModulationEngine.ts", "2h", "test: line draw/dissolve lifecycle", "not_started"),
    ("1", "Morning", "FieldState smoke tests", "Vitest setup", "1h", "test: update + subscribe", "not_started"),
    ("1", "Morning", "ModulationEngine tests", "Vitest setup", "1h", "test: envelope curve, bus output", "not_started"),
    ("2", "Afternoon", "CodeBlock.ts", "monaco-editor", "3h", "test: create/focus/blur/theme", "not_started"),
    ("2", "Afternoon", "BlockManager.ts", "CodeBlock.ts", "3h", "test: CRUD, position, bridge sync", "not_started"),
    ("2", "Afternoon", "Block spawn animation", "BlockManager.ts", "2h", "visual: bridge write trigger", "not_started"),
    ("3", "Evening", "ConversationLayer.ts", "Field.ts", "3h", "test: render + history pruning", "not_started"),
    ("3", "Evening", "Audio layer", "ModulationEngine.ts", "3h", "test: context lifecycle", "not_started"),
    ("3", "Evening", "Camera panning", "Field.ts", "2h", "test: offset tracking", "not_started"),
    ("4", "Rest", "spaceman.svg", "-", "2h", "visual: load as Image", "not_started"),
    ("4", "Rest", "Voice sequencing", "VoiceLayer", "2h", "test: state machine progression", "not_started"),
    ("4", "Rest", "Oval slot wiring", "OvalStadium.ts", "1h", "test: bridge → slot activation", "not_started"),
    ("4", "Rest", "Full ceremony test", "all", "2h", "integration: ground → elevated → return", "not_started"),
]


def write_json(inventory: list[dict], ceremony: list[dict]) -> Path:
    SNAPSHOTS.mkdir(exist_ok=True)
    today = date.today().isoformat()
    out = SNAPSHOTS / f"glass-state-{today}.json"

    payload = {
        "snapshot_date": today,
        "root": str(ROOT),
        "total_files": len([r for r in inventory if r["status"] == "done"]),
        "total_lines": sum(r["lines"] for r in inventory),
        "missing_count": len([r for r in inventory if r["status"] == "not_started"]),
        "inventory": inventory,
        "ceremony_states": ceremony,
        "dev_phases": [
            {"phase": p[0], "cycle": p[1], "component": p[2], "depends_on": p[3],
             "effort": p[4], "test_strategy": p[5], "status": p[6]}
            for p in DEV_PHASES
        ],
    }
    out.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    print(f"  JSON → {out}")
    return out


def write_xlsx(inventory: list[dict], ceremony: list[dict]) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  openpyxl not available — skipping xlsx", file=sys.stderr)
        return Path()

    SNAPSHOTS.mkdir(exist_ok=True)
    today = date.today().isoformat()
    out = SNAPSHOTS / f"glass-state-{today}.xlsx"

    wb = Workbook()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="2B2B2B")
    header_font_white = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    done_fill = PatternFill("solid", fgColor="1A3A1A")
    missing_fill = PatternFill("solid", fgColor="FFFF00")
    blue_font = Font(name="Arial", size=10, color="0000FF")

    def style_header(ws, cols):
        for c, label in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # ── Sheet 1: Component Inventory ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Component Inventory"
    cols1 = ["Component", "Category", "File Path", "Lines", "Status",
             "Planned in DESIGN.md", "Exports", "Notes"]
    style_header(ws1, cols1)

    for i, row in enumerate(inventory, 2):
        ws1.cell(row=i, column=1, value=row["component"]).font = body_font
        ws1.cell(row=i, column=2, value=row["category"]).font = body_font
        ws1.cell(row=i, column=3, value=row["path"]).font = body_font
        lines_cell = ws1.cell(row=i, column=4, value=row["lines"])
        lines_cell.font = blue_font if row["status"] == "done" else body_font
        lines_cell.number_format = "#,##0"
        status_cell = ws1.cell(row=i, column=5, value=row["status"].upper())
        status_cell.font = body_font
        if row["status"] == "done":
            status_cell.fill = done_fill
            status_cell.font = Font(name="Arial", size=10, color="00FF00")
        else:
            status_cell.fill = missing_fill
            status_cell.font = Font(name="Arial", size=10, color="000000")
        ws1.cell(row=i, column=6, value="Yes" if row["designed"] else "Emergent").font = body_font
        ws1.cell(row=i, column=7, value=", ".join(row["exports"])).font = body_font
        ws1.cell(row=i, column=8, value="").font = body_font
        for c in range(1, 9):
            ws1.cell(row=i, column=c).border = border

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 48
    ws1.column_dimensions["D"].width = 8
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 20
    ws1.column_dimensions["G"].width = 40
    ws1.column_dimensions["H"].width = 20

    # summary row
    summary_row = len(inventory) + 2
    ws1.cell(row=summary_row, column=1, value="TOTAL").font = header_font
    ws1.cell(row=summary_row, column=4, value=f"=SUM(D2:D{len(inventory)+1})")
    done_count = len([r for r in inventory if r["status"] == "done"])
    ws1.cell(row=summary_row, column=5, value=f"{done_count}/{len(inventory)} done").font = header_font

    # ── Sheet 2: Ceremony States ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Ceremony States")
    cols2 = ["ThresholdState", "Sustain", "LFO Rate (Hz)", "LFO Depth",
             "Disk Scale", "Oval Opacity", "Voice Alpha", "Visual Description"]
    style_header(ws2, cols2)

    for i, row in enumerate(ceremony, 2):
        ws2.cell(row=i, column=1, value=row["state"]).font = body_font
        ws2.cell(row=i, column=2, value=row["sustain"]).font = blue_font
        ws2.cell(row=i, column=2).number_format = "0.00"
        ws2.cell(row=i, column=3, value=row["lfo_rate"]).font = blue_font
        ws2.cell(row=i, column=3).number_format = "0.00"
        ws2.cell(row=i, column=4, value=row["lfo_depth"]).font = blue_font
        ws2.cell(row=i, column=4).number_format = "0.000"
        ws2.cell(row=i, column=5, value=float(row["disk_scale"])).font = body_font
        ws2.cell(row=i, column=5).number_format = "0.00"
        ws2.cell(row=i, column=6, value=float(row["oval_opacity"])).font = body_font
        ws2.cell(row=i, column=6).number_format = "0.00"
        ws2.cell(row=i, column=7, value=float(row["voice_alpha"])).font = body_font
        ws2.cell(row=i, column=7).number_format = "0.00"
        ws2.cell(row=i, column=8, value=row["description"]).font = body_font
        for c in range(1, 9):
            ws2.cell(row=i, column=c).border = border

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 12
    ws2.column_dimensions["H"].width = 52

    # ── Sheet 3: Development Phases ───────────────────────────────────────────
    ws3 = wb.create_sheet("Development Phases")
    cols3 = ["Phase", "Cycle", "Component", "Depends On", "Est. Effort",
             "Test Strategy", "Status"]
    style_header(ws3, cols3)

    for i, p in enumerate(DEV_PHASES, 2):
        for c, val in enumerate(p, 1):
            cell = ws3.cell(row=i, column=c, value=val)
            cell.font = body_font
            cell.border = border

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 28
    ws3.column_dimensions["D"].width = 22
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 36
    ws3.column_dimensions["G"].width = 14

    wb.save(str(out))
    print(f"  XLSX → {out}")
    return out


def main():
    print(f"Glass build snapshot — {date.today().isoformat()}")
    print(f"  Root: {ROOT}")

    inventory = build_inventory()
    ceremony = read_modulation_data()

    done = [r for r in inventory if r["status"] == "done"]
    missing = [r for r in inventory if r["status"] == "not_started"]
    total_lines = sum(r["lines"] for r in inventory)

    print(f"  Files: {len(done)} done, {len(missing)} not started")
    print(f"  Lines: {total_lines:,}")

    write_json(inventory, ceremony)
    write_xlsx(inventory, ceremony)

    print("Done.")


if __name__ == "__main__":
    main()
